from openpyxl import load_workbook


index = {1:4,50:3,100:2,500:1,1000:0} #used by setUltrasound()


class M13Phage:
        
    def __init__(self, fileName):

        #initialize hashmap 
        self.intensity_to_voltage_dict = {}

        #read the excel file provided by Arash
        workbook = load_workbook(fileName, data_only=True)
        worksheet = workbook.active
    
        #iterate through each row starting at the 5th row
        for row in worksheet.iter_rows(min_row=5): 

            #save the row's column values into a list
            row_values = [cell.value for cell in row]

            #intensity is at column B of the excel file
            intensity = row_values[2] 

            #corresponding voltages are at column C onwards
            voltages = row_values[3:] 

            #add to hashmap
            self.intensity_to_voltage_dict[intensity] = voltages

        #also initialize a list of all possible input ultrasound intensities (helps when interpolating an input)
        self.intensities = list(self.intensity_to_voltage_dict.keys())

        #initialize input intensenty at 0mW/mm^2
        self.input_ultrasound = 0

        #initialzie output source voltage at 0mV
        self.output_voltage = 0


    #function to set the user ultrasound input
    def setUltrasound(self, intensity, num_chan):

        voltage = 0
    
        for i in range(1, len(self.intensities)):
        
            #if exact intensity is in the hashmap, set the output_voltage with a quick lookup
            if self.intensities[i] == intensity:
                
                voltage = self.intensity_to_voltage_dict[intensity][index[num_chan]]
            
            #if the input intensity is between two stored intensities -> interpolate
            elif intensity > self.intensities[i-1] and intensity < self.intensities[i]:

                x2 = self.intensities[i]
                x1 = self.intensities[i-1]

                y2 = self.intensity_to_voltage_dict[self.intensities[i]][index[num_chan]]
                y1 = self.intensity_to_voltage_dict[self.intensities[i-1]][index[num_chan]]
                                       
                voltage = (((y2-y1)/(x2-x1))*(intensity*(intensity-x1))) + y1

        #set object variables
        self.input_ultrasound = intensity
        self.output_voltage = voltage


    #function to get the output voltage
    def getVoltage(self):
        return self.output_voltage








